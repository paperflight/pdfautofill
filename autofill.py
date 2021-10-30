#! /usr/bin/python

import os, sys
import pdfrw


ANNOT_KEY = '/Annots'           # key for all annotations within a page
ANNOT_FIELD_KEY = '/T'          # Name of field. i.e. given ID of field
ANNOT_FORM_type = '/FT'         # Form type (e.g. text/button)
ANNOT_FORM_button = '/Btn'      # ID for buttons, i.e. a checkbox
ANNOT_FORM_text = '/TU'         # ID for detail
SUBTYPE_KEY = '/Subtype'
WIDGET_SUBTYPE_KEY = '/Widget'

data_dict={
    'form[0].#subform[0].Line4_DaytimeTelephoneNumber[0]':'Mouren',
    'page0_Pt1Line2c_MiddleName[0]':'Mail'
}
#data_dict={
#}

from openpyxl import Workbook
def inspect_value(input_pdf_path):
    key_list = []
    template_pdf=pdfrw.PdfReader(input_pdf_path)
    for page_number, page in enumerate(template_pdf.pages):
        annotations=page[ANNOT_KEY]
        for annotation in annotations:
            if annotation[SUBTYPE_KEY]==WIDGET_SUBTYPE_KEY:
                key=annotation[ANNOT_FIELD_KEY][1:-1]
                try:
                    key = str(bytes.fromhex(key).decode('utf-16'))
                except ValueError:
                    key = key.split('.')[-1]
                key = 'page' + str(page_number) + '_' +key
                if annotation[ANNOT_FORM_type] == ANNOT_FORM_button:
                    checkbox_dict = annotation['/AP']['/D']
                    on_key = ''
                    for on_keys in checkbox_dict.keys():
                        if on_keys != '/Off':
                            on_key = on_keys[1::]
                            break
                    print(key, '"' + str(annotation['/V']) + '"', '"' + str(on_key) + '"')
                else:
                    print(key, '"' + str(annotation['/V']) + '"')
                key_list.append(key)
                
def inspect_print(input_pdf_path):
    key_list = []
    template_pdf=pdfrw.PdfReader(input_pdf_path)
    template_pdf.Root.AcroForm.update(pdfrw.PdfDict(NeedAppearances=pdfrw.PdfObject('true')))
    for page_number, page in enumerate(template_pdf.pages):
        annotations=page[ANNOT_KEY]
        for annotation in annotations:
            if annotation[SUBTYPE_KEY]==WIDGET_SUBTYPE_KEY:
                key=annotation[ANNOT_FIELD_KEY][1:-1]
                try:
                    key = str(bytes.fromhex(key).decode('utf-16'))
                except ValueError:
                    key = key.split('.')[-1]
                if annotation[ANNOT_FORM_type] == ANNOT_FORM_button:
                    checkbox_dict = annotation['/AP']['/D']
                    on_key = ''
                    for on_keys in checkbox_dict.keys():
                        if on_keys != '/Off':
                            on_key = on_keys[1::]
                            break
                    annotation.update(
                        pdfrw.PdfDict(V=pdfrw.PdfName(on_key), AS=pdfrw.PdfName(on_key)) # default checkbox value is 'Off'
                    )
                else:
                    annotation.update(
                        pdfrw.PdfDict(V=pdfrw.PdfName(key)) # default checkbox value is 'Off'
                    )
    pdfrw.PdfWriter().write(input_pdf_path[0:-4] + '-investigate.pdf',template_pdf)
        
def inspect(input_pdf_path, input_excel_path=None):
    key_list = []
    text_list = []
    template_pdf=pdfrw.PdfReader(input_pdf_path)
    template_pdf.Root.AcroForm.update(pdfrw.PdfDict(NeedAppearances=pdfrw.PdfObject('true')))
    for page_number, page in enumerate(template_pdf.pages):
        annotations=page[ANNOT_KEY]
        for annotation in annotations:
            if annotation[SUBTYPE_KEY]==WIDGET_SUBTYPE_KEY:
                key=annotation[ANNOT_FIELD_KEY][1:-1]
                try:
                    key = str(bytes.fromhex(key).decode('utf-16'))
                except ValueError:
                    key = key.split('.')[-1]
                key = 'page' + str(page_number) + '_' +key
                print(key)
                key_list.append(key)
                if annotation[ANNOT_FORM_text] is not None:
                    text_list.append(annotation[ANNOT_FORM_text])
                else:
                    text_list.append('')
    if input_excel_path is None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = os.path.basename(input_pdf_path)[0:-4]
        for key_row, (key, text) in enumerate(zip(key_list, text_list)):
            sheet['A'+str(key_row + 1)] = key
            sheet['C'+str(key_row + 1)] = text
        print('Data write to sheet ' + sheet.title)
        workbook.save(filename = input_pdf_path[0:-4] + '.xlsx')
    else:
        workbook = load_workbook(input_excel_path)
        sheet_name = os.path.basename(input_pdf_path)[0:-4]
        sheet = workbook.create_sheet(title=sheet_name)
        for key_row, (key, text) in enumerate(zip(key_list, text_list)):
            sheet['A'+str(key_row + 1)] = key
            sheet['C'+str(key_row + 1)] = text
        print('Data write to sheet ' + sheet.title)
        workbook.save(filename = input_excel_path)


def write_fillable_pdf(input_pdf_path,output_pdf_path,data_dict):
    if not os.path.isfile(input_pdf_path):
        print('Could not find ' + input_pdf_path + '. Skip.')
        return
    template_pdf=pdfrw.PdfReader(input_pdf_path)
    template_pdf.Root.AcroForm.update(pdfrw.PdfDict(NeedAppearances=pdfrw.PdfObject('true')))
    for page_number, page in enumerate(template_pdf.pages):
        annotations=page[ANNOT_KEY]
        for annotation in annotations:
            if annotation[SUBTYPE_KEY]==WIDGET_SUBTYPE_KEY:
                key=annotation[ANNOT_FIELD_KEY][1:-1]
                try:
                    key = str(bytes.fromhex(key).decode('utf-16'))
                except ValueError:
                    key = key.split('.')[-1]
                key = 'page' + str(page_number) + '_' +key
                try:
                    if data_dict[key] == '':
                        continue
                    if annotation[ANNOT_FORM_type] == ANNOT_FORM_button:
                        checkbox_dict = annotation['/AP']['/D']
                        on_key = ''
                        for on_keys in checkbox_dict.keys():
                            if on_keys != '/Off':
                                on_key = on_keys[1::]
                                print(on_key)
                                break
                        if 'Yes' in data_dict[key] or 'yes' in data_dict[key] or 'On' in data_dict[key] or on_key == data_dict[key]:
                            annotation.update(
                                pdfrw.PdfDict(V=pdfrw.PdfName(on_key), AS=pdfrw.PdfName(on_key)) # default checkbox value is 'Off'
                            )
                        else:
                             annotation.update(
                                pdfrw.PdfDict(V=pdfrw.PdfName('Off'), AS=pdfrw.PdfName('Off')) # default checkbox value is 'Off'
                            )
                    else:
                        annotation.update(
                            pdfrw.PdfDict(V='{}'.format(data_dict[key]))
                        )
                except KeyError:
                    print('Missing in: ' + input_pdf_path)
                    print('Missing: ' + key)
                        
    pdfrw.PdfWriter().write(output_pdf_path,template_pdf)
    
import pikepdf
def decrpt(input_pdf_path):
    pdf = pikepdf.open(input_pdf_path, allow_overwriting_input=True)
    pdf.save(input_pdf_path)
    print('Decrpyt Complete')


from openpyxl import load_workbook
def read_excel(input_excel_path):
    workbook = load_workbook(filename=input_excel_path)
    sheet = workbook.active
    print('Extracting from ' + sheet.title)
    for data in sheet.iter_rows(values_only=True):
        print(data[0], data[1])
        data_dict[data[0]] = data[1]

def run_all(input_excel_path):
    path = os.getcwd() + '/'
    workbook = load_workbook(filename=input_excel_path, data_only=True)
    for sheet_name_index, sheet_name in enumerate(workbook.sheetnames):
        data_dict = {}
        sheet = workbook.worksheets[sheet_name_index]
        print('Extracting from ' + sheet.title)
        for data in sheet.iter_rows(min_col=1, max_col=2, values_only=True):
            key = data[0]
            if data[0] is not None and len(data[0].split('.')) > 1:
                page_number = data[0].split('.')[-2]
                key = 'page' + str(page_number[-2]) + '_' + data[0].split('.')[-1]
            if data[1] is None:
                print(str(key), '')
                data_dict[str(key)] = ''
            else:
                print(key, data[1])
                data_dict[str(key)] = str(data[1])
        write_fillable_pdf(path + sheet_name + '.pdf', path + sheet_name + '-fill.pdf', data_dict)
    
    
if __name__ == '__main__':
    if sys.argv[1] == 'inspect':
        if len(sys.argv) == 3:
            inspect(sys.argv[2])
        elif len(sys.argv) == 4:
            inspect(sys.argv[2], sys.argv[3])
    elif sys.argv[1] == 'inspect_value':
        inspect_value(sys.argv[2])
    elif sys.argv[1] == 'inspect_print':
        inspect_print(sys.argv[2])
    elif sys.argv[1] == 'decrpt':
        decrpt(sys.argv[2])
    elif sys.argv[1] == 'write':
        write_fillable_pdf(sys.argv[2], sys.argv[3], data_dict)
    elif sys.argv[1] == 'read_excel':
        read_excel(sys.argv[2])
    else:
        run_all(sys.argv[1])
